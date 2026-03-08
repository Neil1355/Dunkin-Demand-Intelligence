from openpyxl import load_workbook

file_path = r"C:\\Users\\neilb\\Downloads\\dunkin_export_2026-03-08.xlsx"
wb = load_workbook(file_path, data_only=False)
ws = wb.active

print('Title:', ws.title)
print('Max row/col:', ws.max_row, ws.max_column)

# Print first 20 rows x first 18 cols
for r in range(1, 21):
    vals = []
    for c in range(1, 19):
        v = ws.cell(r, c).value
        vals.append(v)
    print(r, vals)

# Count non-empty AM/PM cells in B:O
non_empty = 0
for r in range(1, ws.max_row + 1):
    for c in range(2, 16):
        if ws.cell(r, c).value not in (None, ""):
            non_empty += 1
print('Non-empty B:O cells:', non_empty)

# Print header date row and number format info
for c in range(1, 16):
    cell = ws.cell(2, c)
    if cell.value is not None:
        print('R2C', c, 'value=', cell.value, 'fmt=', cell.number_format)
