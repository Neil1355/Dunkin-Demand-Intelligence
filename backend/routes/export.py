from flask import Blueprint, send_file
import pandas as pd
from openpyxl.styles import PatternFill
from ..models.db import get_connection

export_bp = Blueprint("export", __name__)

@export_bp.get("/forecast")
def export_forecast():
    conn = get_connection()
    df = pd.read_sql("""
        SELECT
            p.product_name,
            fh.target_date,
            fh.predicted_quantity
        FROM forecast_history fh
        JOIN products p ON fh.product_id = p.product_id
        WHERE fh.forecast_date = CURRENT_DATE
    """, conn)

    path = "/tmp/forecast.xlsx"
    df.to_excel(path, index=False)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active

    green = PatternFill("solid", fgColor="C6EFCE")

    for row in ws.iter_rows(min_row=2):
        row[2].fill = green

    wb.save(path)

    return send_file(path, as_attachment=True)
