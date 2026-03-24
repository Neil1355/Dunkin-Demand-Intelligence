from flask import Blueprint, request, send_file, jsonify
import pandas as pd
from datetime import timedelta, datetime
from models.db import get_connection, return_connection
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill
from typing import cast
import tempfile
import os

throwaway_export_bp = Blueprint("throwaway_export", __name__, url_prefix="/throwaway")

@throwaway_export_bp.get("/export")
def export_throwaway():
    """
    Export weekly data in AM/PM format matching the Dunkin template exactly
    - Row 1: Empty
    - Row 2: DATE: and date value
    - Row 3: Day names (SUN, MON, TUE, WED, THU, FRI, SAT)
    - Row 4: AM/PM headers
    - Rows 5+: Products organized by category
    - Bottom rows: Donuts Bought, Donuts Sold (renamed to Difference), Throwaway
    """
    try:
        store_id = request.args.get("store_id", type=int)
        week_start = request.args.get("week_start")

        if not store_id:
            return jsonify({"error": "store_id required"}), 400
        
        # Default to current week's Sunday
        if not week_start:
            today = datetime.now().date()
            days_since_sunday = (today.weekday() + 1) % 7
            week_start = today - timedelta(days=days_since_sunday)
        else:
            week_start = pd.to_datetime(week_start).date()
        
        dates = [week_start + timedelta(days=i) for i in range(7)]

        conn = get_connection()
        cur = conn.cursor()

        # Fetch all active products
        cur.execute("""
            SELECT product_id, product_name, product_type
            FROM products
            WHERE is_active = TRUE
            ORDER BY 
                CASE product_type
                    WHEN 'croissant' THEN 1
                    WHEN 'bagel' THEN 2
                    WHEN 'donut' THEN 3
                    WHEN 'muffin' THEN 4
                    WHEN 'munchkin' THEN 5
                    ELSE 6
                END,
                product_name
        """)
        products = cur.fetchall()

        # Fetch throwaway data and products with activity in the selected week
        cur.execute("""
            SELECT 
                p.product_name,
                p.product_type,
                dt.date,
                COALESCE(dt.produced, 0) AS produced,
                COALESCE(dt.waste, 0) AS waste
            FROM products p
            INNER JOIN daily_throwaway dt ON p.product_id = dt.product_id
                AND dt.store_id = %s
                AND dt.date BETWEEN %s AND %s
            WHERE p.is_active = TRUE
              AND (dt.produced > 0 OR dt.waste > 0)
            ORDER BY 
                CASE p.product_type
                    WHEN 'croissant' THEN 1
                    WHEN 'bagel' THEN 2
                    WHEN 'donut' THEN 3
                    WHEN 'muffin' THEN 4
                    WHEN 'munchkin' THEN 5
                    ELSE 6
                END,
                p.product_name, dt.date
        """, (store_id, dates[0], dates[-1]))
        
        rows = cur.fetchall()
        cur.close()
        return_connection(conn)

        # Build data structure: {product_name: [am_sun, pm_sun, am_mon, pm_mon, ...]}
        # Only include products that have data in this week (from rows)
        product_data = {}
        product_types = {}
        active_products_this_week = set()
        
        # First pass: identify which products have data
        for row in rows:
            active_products_this_week.add(row['product_name'])
        
        # Only initialize products that are active this week
        for product in products:
            product_name = product['product_name']
            if product_name in active_products_this_week:
                product_data[product_name] = [None] * 14
                product_types[product_name] = product['product_type']

        # Fill in actual data
        for row in rows:
            product_name = row['product_name']
            date = row['date']
            produced = row['produced']
            waste = row['waste']
            
            if product_name in product_data and date:
                try:
                    day_index = dates.index(date)
                    product_data[product_name][day_index * 2] = int(produced) if produced else None
                    product_data[product_name][day_index * 2 + 1] = int(waste) if waste else None
                except (ValueError, IndexError):
                    continue

        # Create Excel workbook
        wb = Workbook()
        ws_active = wb.active
        if ws_active is None:
            return jsonify({"error": "Failed to initialize workbook worksheet"}), 500
        ws = cast(Worksheet, ws_active)
        ws.title = "Throwaway"

        current_row = 1

        # Row 1: Empty
        current_row += 1

        # Row 2: DATE: with all 7 dates
        ws.cell(current_row, 1, "DATE:")
        for i, date in enumerate(dates):
            ws.cell(current_row, 2 + i * 2, date)
            ws.cell(current_row, 2 + i * 2).number_format = 'mm/dd/yyyy'
        current_row += 1

        # Row 3: Day names (SUN, MON, etc.) + PM TTL header
        day_names = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]
        for i, day in enumerate(day_names):
            ws.cell(current_row, 2 + i * 2, day)
        ws.cell(current_row, 16, "PM TTL")
        current_row += 1

        # Row 4: AM/PM headers
        for i in range(7):
            ws.cell(current_row, 2 + i * 2, "AM")
            ws.cell(current_row, 3 + i * 2, "PM")
        current_row += 1

        # Track totals for donuts and munchkins separately
        donut_totals_produced = [0] * 7
        donut_totals_waste = [0] * 7
        munchkin_totals_produced = [0] * 7
        munchkin_totals_waste = [0] * 7

        # Smart grouping: reclassify "other" products based on keywords
        def get_smart_category(product):
            product_type = product['product_type']
            product_name_lower = product['product_name'].lower()
            
            # If already categorized, keep it unless it's 'other'
            if product_type != 'other':
                return product_type
            
            # Smart classification for 'other' items
            if 'munchkin' in product_name_lower or 'munch' in product_name_lower:
                return 'munchkin'
            elif any(keyword in product_name_lower for keyword in ['donut', 'glazed', 'frosted', 'chocolate', 'jelly', 'bavarian', 'cruller', 'filled']):
                return 'donut'
            else:
                return 'other'
        
        # Group products by category
        categories = [
            ('croissant', 'Plain Croissants'),
            ('bagel', 'Bagels'),
            ('donut', 'Donuts'),
            ('muffin', 'Muffins'),
            ('munchkin', 'Munchkins'),
            ('other', 'Other Items')
        ]

        for product_type, category_label in categories:
            # Get products in this category (with smart grouping)
            category_products = [
                p for p in products
                if get_smart_category(p) == product_type and p['product_name'] in active_products_this_week
            ]
            
            if not category_products:
                continue

            # Add category header
            ws.cell(current_row, 1, category_label)
            ws.cell(current_row, 1).font = Font(bold=True)
            current_row += 1

            # Add products
            for product in category_products:
                product_name = product['product_name']
                values = product_data.get(product_name, [None] * 14)
                
                ws.cell(current_row, 1, product_name)
                
                # Fill in AM/PM values and calculate PM total
                pm_total = 0
                for i, val in enumerate(values):
                    if val is not None and val != 0:
                        ws.cell(current_row, 2 + i, val)
                        # If it's a PM column (odd index), add to PM total
                        if i % 2 == 1:
                            pm_total += val
                
                # Add PM TTL column
                if pm_total > 0:
                    ws.cell(current_row, 16, pm_total)
                
                # Accumulate totals for donuts and munchkins (use smart category)
                for day in range(7):
                    produced = values[day * 2] if values[day * 2] else 0
                    waste = values[day * 2 + 1] if values[day * 2 + 1] else 0
                    
                    smart_category = get_smart_category(product)
                    if smart_category == 'donut':
                        donut_totals_produced[day] += produced
                        donut_totals_waste[day] += waste
                    elif smart_category == 'munchkin':
                        munchkin_totals_produced[day] += produced
                        munchkin_totals_waste[day] += waste
                
                current_row += 1

            # Add blank row after category
            current_row += 1

        # Add summary rows for Donuts
        current_row += 1
        
        # Donuts Bought (produced in AM columns only)
        ws.cell(current_row, 1, "Donuts Bought")
        ws.cell(current_row, 1).font = Font(bold=True)
        for day in range(7):
            if donut_totals_produced[day] > 0:
                ws.cell(current_row, 2 + day * 2, donut_totals_produced[day])
        ws.cell(current_row, 16, sum(donut_totals_produced))
        current_row += 1

        # Donuts Sold (produced - waste in AM columns only)
        ws.cell(current_row, 1, "Donuts Sold")
        ws.cell(current_row, 1).font = Font(bold=True)
        total_sold = 0
        for day in range(7):
            sold = donut_totals_produced[day] - donut_totals_waste[day]
            if sold > 0:
                ws.cell(current_row, 2 + day * 2, sold)
                total_sold += sold
        ws.cell(current_row, 16, total_sold)
        current_row += 1

        # Difference (Bought - Sold = waste, shown in AM columns)
        ws.cell(current_row, 1, "Difference")
        ws.cell(current_row, 1).font = Font(bold=True)
        for day in range(7):
            diff = donut_totals_waste[day]
            if diff > 0:
                ws.cell(current_row, 2 + day * 2, diff)
        ws.cell(current_row, 16, sum(donut_totals_waste))
        current_row += 1

        # Throwaway (waste in PM columns only)
        ws.cell(current_row, 1, "Throwaway")
        ws.cell(current_row, 1).font = Font(bold=True)
        for day in range(7):
            if donut_totals_waste[day] > 0:
                ws.cell(current_row, 3 + day * 2, donut_totals_waste[day])
        current_row += 1

        # Empty rows
        current_row += 1
        current_row += 1
        current_row += 1

        # Tot PM Donut Throw
        ws.cell(current_row, 1, "Tot PM Donut Throw")
        ws.cell(current_row, 1).font = Font(bold=True)
        total_pm_throw = sum(donut_totals_waste)
        ws.cell(current_row, 2, total_pm_throw)
        # Also show in each PM column
        for day in range(7):
            if donut_totals_waste[day] > 0:
                ws.cell(current_row, 3 + day * 2, donut_totals_waste[day])
        current_row += 1

        # Add summary rows for Munchkins
        current_row += 1
        current_row += 1
        
        # Second set of day headers forMunchkins section
        for i, day in enumerate(day_names):
            ws.cell(current_row, 2 + i * 2, day)
        ws.cell(current_row, 16, "PM TTL")
        current_row += 1
        
        # AM/PM headers again
        for i in range(7):
            ws.cell(current_row, 2 + i * 2, "AM")
            ws.cell(current_row, 3 + i * 2, "PM")
        current_row += 1
        
        # Munchkins Bought
        ws.cell(current_row, 1, "Muffins")
        ws.cell(current_row, 1).font = Font(bold=True)
        current_row += 1
        
        # Add muffin products if any (empty for now based on screenshot)
        current_row += 1
        current_row += 1
        
        # Munchkins section
        ws.cell(current_row, 1, "Munchkins")
        ws.cell(current_row, 1).font = Font(bold=True)
        current_row += 1
        
        # Add blank rows
        current_row += 1
        current_row += 1
        current_row += 1
        
        # Munchkins Bought
        ws.cell(current_row, 1, "Munchkins Bought")
        ws.cell(current_row, 1).font = Font(bold=True)
        for day in range(7):
            if munchkin_totals_produced[day] > 0:
                ws.cell(current_row, 2 + day * 2, munchkin_totals_produced[day])
        ws.cell(current_row, 16, sum(munchkin_totals_produced))
        current_row += 1

        # Munchkins Sold
        ws.cell(current_row, 1, "Munchkins Sold")
        ws.cell(current_row, 1).font = Font(bold=True)
        total_munchkin_sold = 0
        for day in range(7):
            sold = munchkin_totals_produced[day] - munchkin_totals_waste[day]
            if sold > 0:
                ws.cell(current_row, 2 + day * 2, sold)
                total_munchkin_sold += sold
        ws.cell(current_row, 16, total_munchkin_sold)
        current_row += 1
        
        # Munchkin Difference
        ws.cell(current_row, 1, "Difference")
        ws.cell(current_row, 1).font = Font(bold=True)
        for day in range(7):
            diff = munchkin_totals_waste[day]
            if diff > 0:
                ws.cell(current_row, 2 + day * 2, diff)
        ws.cell(current_row, 16, sum(munchkin_totals_waste))
        current_row += 1

        # Munchkin Throwaway
        ws.cell(current_row, 1, "Throwaway")
        ws.cell(current_row, 1).font = Font(bold=True)
        for day in range(7):
            if munchkin_totals_waste[day] > 0:
                ws.cell(current_row, 3 + day * 2, munchkin_totals_waste[day])
        current_row += 1

        # Improve sheet readability so date cells don't render as #### in Excel.
        ws.column_dimensions['A'].width = 32
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            ws.column_dimensions[col].width = 11
        ws.column_dimensions['P'].width = 10

        # Save to temporary file
        temp_dir = tempfile.gettempdir()
        filename = f"Dunkin_Throwaways_{week_start.strftime('%m.%d.%y')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        wb.save(filepath)

        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error exporting throwaway data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
